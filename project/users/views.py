#################
#### imports ####
#################

from flask import ( flash, redirect, render_template, request,
                    url_for, Blueprint,current_app)
from flask_login import login_user, login_required, logout_user
from .form import LoginForm, RegisterForm, UploadForm
from project.models import User, bcrypt
from project import db
from werkzeug import secure_filename
from openpyxl import load_workbook
import simplejson as json
import os
################
#### config ####
################

users_blueprint = Blueprint(
    'users', __name__,
    template_folder='templates'
)

################
#### routes ####
################

@users_blueprint.route('/login',methods=['GET', 'POST'])
def login():
    error = None
    form = LoginForm(request.form)
    if request.method == 'POST':
        if form.validate_on_submit():
            user = User.query.filter_by(name=request.form['username']).first()
            if user is not None and bcrypt.check_password_hash(
                user.password, request.form['password']
            ):
                # session['logged_in'] = True
                login_user(user)
                flash('You were logged in.')
                return redirect(url_for('home.dashboard'))
            else:
                error = 'Invalid username or password.'
    return render_template('login.html',form=form, error=error)


@users_blueprint.route('/logout')
@login_required
def logout():
    logout_user()
    # session.pop('logged_in',None)
    flash('You were logged out.')
    return redirect(url_for('home.welcome'))


@users_blueprint.route('/register', methods=['GET', 'POST'])   # pragma: no cover
def register():
    error = None
    form = RegisterForm()
    if form.validate_on_submit():
        user = User(
            name=form.username.data,
            email=form.email.data,
            password=form.password.data
        )
        db.session.add(user)
        db.session.commit()
        login_user(user)
        return redirect(url_for('home.dashboard'))
    return render_template('register.html', form=form)

@users_blueprint.route('/uploader', methods=['GET','POST'])
def uploader():
    error = None
    form = UploadForm()
    if form.validate_on_submit():
        # print(current_app.config['UPLOAD_JSON'])
        f = form.file.data
        filename = secure_filename(f.filename)
        # print(f)

        wb = load_workbook(f)

        jsonArray = []

        for sheet in wb.sheetnames:
            ws = wb[sheet]

            tempDict = {}
            header = [cell.value.lower() for cell in ws[1]]

            tempDict['subject']=sheet.lower()
            jsonArray.append(tempDict)

            for row in ws.iter_rows(min_row=2, values_only=True):
                for i,cell in enumerate(row):
                    tempDict = {}
                    tempDict[header[i]]=cell.lower()
                    jsonArray.append(tempDict)


        #print(jsonArray)
        jsonPath = current_app.config['UPLOAD_JSON']
        j = json.dumps(jsonArray,indent=4,sort_keys=True)
        newfilename = os.path.splitext(filename)[0] + '.json'
        p = os.path.join(jsonPath, newfilename)
        with open(p, "w") as fn:
            fn.write(j)

        return 'buh'
        #return app.instance_path +"\n"+app.root_path
    else:
        print(form.errors)
        return render_template('uploader.html', form=form, error=error)
