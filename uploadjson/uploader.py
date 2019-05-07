from flask import Flask, render_template, request
from werkzeug import secure_filename
from openpyxl import load_workbook
import simplejson as json
import os
app = Flask(__name__)


def fixName(name):
    name = os.path.splitext(name)[0]
    return name+'.json'

@app.route('/uploader', methods = ['GET','POST'])
def upload():
    if request.method == 'POST':
        f = request.files['file']         
        filename = secure_filename(f.filename)
        print(f)

        wb = load_workbook(f)

        jsonArray = []

        for sheet in wb.sheetnames:
            ws = wb[sheet]

            tempDict = {}
            header = [cell.value.lower() for cell in ws[1]]

            tempDict['subject']=sheet.lower()
            jsonArray.append(tempDict)
    
            for row in ws.iter_rows(min_row=2, values_only=True):
                for i,cell in enumerate(row):
                    tempDict = {}
                    tempDict[header[i]]=cell.lower()
                    jsonArray.append(tempDict)

 
        #print(jsonArray)
        j = json.dumps(jsonArray,indent=4,sort_keys=True)      
        newfilename = fixName(filename)
        p = os.path.join(app.root_path, newfilename)
        with open(p, "w") as fn:
            fn.write(j)
        
        return p
        #return app.instance_path +"\n"+app.root_path
    else:
        return "Not posted"



if __name__ == "__main__":
    app.run(debug=True)