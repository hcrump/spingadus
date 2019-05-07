from openpyxl import load_workbook
import simplejson as json


wb = load_workbook('./small.xlsx')

jsonArray = []

for sheet in wb.sheetnames:
    ws = wb[sheet]

    tempDict = {}
    header = [cell.value.lower() for cell in ws[1]]

    tempDict['subject']=sheet.lower()
    jsonArray.append(tempDict)
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        for i,cell in enumerate(row):
            tempDict = {}
            tempDict[header[i]]=cell.lower()
            jsonArray.append(tempDict)

 
#print(jsonArray)
j = json.dumps(jsonArray,indent=4,sort_keys=True)      
print(j)

with open('test.json', 'w') as f:
    f.write(j)
    
            
    